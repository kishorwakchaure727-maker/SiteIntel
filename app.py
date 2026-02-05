import streamlit as st
import pandas as pd
import requests
import re
import time
import hashlib
from bs4 import BeautifulSoup
from io import BytesIO
from unidecode import unidecode
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# -------------------------------------------------
# CONFIG
# -------------------------------------------------
HEADERS = {"User-Agent": "Mozilla/5.0"}
REQUEST_DELAY = 0.15  # Google Maps throttle

try:
    GOOGLE_MAPS_API_KEY = st.secrets.get("GOOGLE_MAPS_API_KEY", "")
except:
    GOOGLE_MAPS_API_KEY = ""

SHORT_FORMS = {
    "RD": "ROAD", "ST": "STREET", "AVE": "AVENUE",
    "BLVD": "BOULEVARD", "DR": "DRIVE",
    "LN": "LANE", "PL": "PLACE"
}

STANDARD_COUNTRIES = {
    "USA": "UNITED STATES OF AMERICA",
    "US": "UNITED STATES OF AMERICA",
    "UNITED STATES": "UNITED STATES OF AMERICA",
    "UK": "UNITED KINGDOM OF GREAT BRITAIN AND NORTHERN IRELAND",
    "UNITED KINGDOM": "UNITED KINGDOM OF GREAT BRITAIN AND NORTHERN IRELAND"
}

# -------------------------------------------------
# HELPERS
# -------------------------------------------------
def normalize_text(text: str) -> str:
    text = unidecode(text).upper()
    text = re.sub(r"[^\w\s]", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()

def hash_address(addr: dict) -> str:
    key = "|".join([
        addr.get("STREET ADDRESS 1", ""),
        addr.get("CITY", ""),
        addr.get("STATE", ""),
        addr.get("PIN CODE", ""),
        addr.get("COUNTRY", "")
    ])
    return hashlib.md5(key.encode()).hexdigest()

def standardize_address(raw: str) -> str:
    """Lightweight top-level standardizer returning primary street line.

    This keeps compatibility with older UI code that expects a string.
    """
    if not raw:
        return ""
    raw_norm = normalize_text(raw)
    for k, v in SHORT_FORMS.items():
        raw_norm = re.sub(rf"\b{k}\b", v, raw_norm, flags=re.I)
    parts = [p.strip() for p in re.split(r",|;|\n", raw_norm) if p.strip()]
    return parts[0] if parts else raw_norm

def enrich_google_maps(rec: dict) -> dict:
    """Placeholder enrichment wrapper used by the UI.

    If you want Google enrichment later, extend this to call the Geocoding API.
    For now it returns the record unchanged to avoid NameError.
    """
    return rec


# -------------------------------
# Dict-style extraction + standardization (top-level)
# -------------------------------
def ensure_scheme(url: str) -> str:
    if not url:
        return ""
    url = url.strip()
    if not url.startswith("http"):
        return "https://" + url.lstrip("/")
    return url


STREET_KEYWORDS = r"\b(STREET|ST\.|ROAD|RD\.|AVE|AVENUE|BOULEVARD|BLVD|DR|DRIVE|LANE|LN|WAY|TERRACE|PLAZA|PL|COURT|CT)\b"
EXCLUDE_SALES_KEYWORDS = [
    "store", "stores", "location", "locations", "dealer", "retail",
    "shop", "franchise", "outlet", "distributor", "sales"
]

# Pages and keywords we should prioritize when hunting for corporate addresses
PREFERRED_PAGE_KEYWORDS = [
    "contact", "contact-us", "contactus", "about", "about-us", "aboutus",
    "head", "head-office", "headquarters", "hq", "office", "offices",
    "locations", "location", "plant", "plants", "manufacturing", "factory", "site", "facility"
]

CANDIDATE_PATHS = [
    "/contact", "/contact-us", "/about", "/about-us", "/locations",
    "/location", "/offices", "/head-office", "/headquarters", "/hq",
    "/plants", "/manufacturing", "/factory", "/site", "/facility"
]


def find_pages_from_home(home_url: str, max_pages=10):
    home = ensure_scheme(home_url)
    pages = [home]
    try:
        r = requests.get(home, headers=HEADERS, timeout=6)
        soup = BeautifulSoup(r.text, "html.parser")
        for a in soup.find_all("a", href=True):
            href = a["href"].strip()
            if href.startswith("/"):
                pages.append(home.rstrip("/") + href)
            elif href.startswith("http"):
                pages.append(href)
            if len(set(pages)) >= max_pages:
                break
    except Exception:
        pass

    base = re.sub(r"/+$", "", home)
    # add preferred candidate paths first
    for p in CANDIDATE_PATHS:
        pages.append(base + p)

    # dedupe and prioritize pages whose path contains preferred keywords
    ordered = []
    seen = set()
    # First, add pages that match preferred keywords
    for p in pages:
        lp = p.lower()
        if any(k in lp for k in PREFERRED_PAGE_KEYWORDS):
            if p not in seen:
                ordered.append(p)
                seen.add(p)
    # Then add the rest preserving order
    for p in pages:
        if p not in seen:
            ordered.append(p)
            seen.add(p)

    return ordered[:max_pages]


def extract_address_site(website: str, prefer_hq: bool = True):
    """Return (raw_address, found_page) or ("", "").

    If `prefer_hq` is True, pages or sections that look like retail/store locations
    will be deprioritized or skipped.
    """
    if not website:
        return "", ""
    domain = re.sub(r"https?://", "", website).split("/", 1)[0]
    pages = find_pages_from_home(website, max_pages=12)

    visited = set()
    # Depth-first crawl (pages + one-level internal links)
    for p in pages:
        try:
            if p in visited:
                continue
            visited.add(p)
            r = requests.get(ensure_scheme(p), headers=HEADERS, timeout=6)
            soup = BeautifulSoup(r.text, "html.parser")
            addr_tag = soup.find("address")
            if addr_tag:
                txt = addr_tag.get_text(" ", strip=True)
                if txt:
                    # prefer corporate pages: boost pages with 'contact'/'about' or 'head office'
                    txt_low = txt.lower()
                    if prefer_hq and any(k in txt_low for k in EXCLUDE_SALES_KEYWORDS):
                        # skip sales pages if preferring HQ
                        pass
                    else:
                        return normalize_text(txt), ensure_scheme(p)

            text = soup.get_text(" ", strip=True)
            # search for strict address candidates
            for line in text.splitlines():
                if len(line) < 10:
                    continue
                cand = line.strip()
                if is_strict_address_candidate(cand):
                    cand_low = cand.lower()
                    if prefer_hq and any(k in cand_low for k in EXCLUDE_SALES_KEYWORDS):
                        # deprioritize store/location blocks
                        continue
                    return normalize_text(cand), ensure_scheme(p)
            # if not found on this page, collect internal links for one more depth
            internal_links = []
            for a in soup.find_all("a", href=True):
                href = a["href"].strip()
                if href.startswith("/"):
                    internal_links.append(ensure_scheme(domain + href))
                elif href.startswith("http") and domain in href:
                    internal_links.append(href)
            # crawl one level of internal links
            for link in internal_links[:6]:
                if link in visited:
                    continue
                visited.add(link)
                try:
                    r2 = requests.get(link, headers=HEADERS, timeout=6)
                    txt2 = BeautifulSoup(r2.text, "html.parser").get_text(" ", strip=True)
                    for line in txt2.splitlines():
                        if is_strict_address_candidate(line):
                            cand_low = line.lower()
                            if prefer_hq and any(k in cand_low for k in EXCLUDE_SALES_KEYWORDS):
                                continue
                            return normalize_text(line), link
                except Exception:
                    continue
        except Exception:
            continue

    # fallback: try duckduckgo-lite search
    try:
        url = "https://html.duckduckgo.com/html/"
        # broaden search terms to prefer contact/location/headquarter/manufacturing pages
        query_terms = "contact OR contact us OR locations OR headquarters OR head office OR plant OR manufacturing OR office"
        q = f"site:{domain} {query_terms}"
        res = requests.post(url, data={"q": q}, headers=HEADERS, timeout=6)
        soup = BeautifulSoup(res.text, "html.parser")
        links = []
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if href.startswith("http") and domain in href:
                links.append(href)
            if len(links) >= 6:
                break
        for link in links:
            try:
                r = requests.get(link, headers=HEADERS, timeout=6)
                txt = BeautifulSoup(r.text, "html.parser").get_text(" ", strip=True)
                for line in txt.splitlines():
                    if is_strict_address_candidate(line):
                        cand_low = line.lower()
                        if prefer_hq and any(k in cand_low for k in EXCLUDE_SALES_KEYWORDS):
                            continue
                        return normalize_text(line), link
            except Exception:
                continue
    except Exception:
        pass

    return "", ""


def extract_all_addresses_site(website: str, limit: int = 20):
    """Return list of (raw_address, found_page) found across the site (breadth-first).

    This is used when the user wants multiple facility/store locations for a company.
    """
    out = []
    if not website:
        return out
    domain = re.sub(r"https?://", "", website).split("/", 1)[0]
    pages = find_pages_from_home(website, max_pages=30)
    visited = set()

    def collect_from_text(text, page):
        for line in text.splitlines():
            cand = line.strip()
            if not cand:
                continue
            if is_strict_address_candidate(cand):
                norm = normalize_text(cand)
                if norm and all(norm != e[0] for e in out):
                    out.append((norm, page))
                    if len(out) >= limit:
                        return True
        return False

    # scan listed pages and one-level internal links
    for p in pages:
        if p in visited:
            continue
        visited.add(p)
        try:
            r = requests.get(ensure_scheme(p), headers=HEADERS, timeout=6)
            soup = BeautifulSoup(r.text, "html.parser")
            # address tags
            for tag in soup.find_all("address"):
                txt = tag.get_text(" ", strip=True)
                if txt and collect_from_text(txt, ensure_scheme(p)):
                    return out

            text = soup.get_text(" ", strip=True)
            if collect_from_text(text, ensure_scheme(p)):
                return out

            # gather internal links to scan shallowly
            links = []
            for a in soup.find_all("a", href=True):
                href = a["href"].strip()
                if href.startswith("/"):
                    links.append(ensure_scheme(domain + href))
                elif href.startswith("http") and domain in href:
                    links.append(href)
            for link in links[:10]:
                if link in visited:
                    continue
                visited.add(link)
                try:
                    r2 = requests.get(link, headers=HEADERS, timeout=6)
                    txt2 = BeautifulSoup(r2.text, "html.parser").get_text(" ", strip=True)
                    if collect_from_text(txt2, link):
                        return out
                except Exception:
                    continue
        except Exception:
            continue

    # fallback DDG
    try:
        url = "https://html.duckduckgo.com/html/"
        q = f"site:{domain} contact address"
        res = requests.post(url, data={"q": q}, headers=HEADERS, timeout=6)
        soup = BeautifulSoup(res.text, "html.parser")
        links = []
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if href.startswith("http") and domain in href:
                links.append(href)
            if len(links) >= 20:
                break
        for link in links:
            try:
                r = requests.get(link, headers=HEADERS, timeout=6)
                txt = BeautifulSoup(r.text, "html.parser").get_text(" ", strip=True)
                if collect_from_text(txt, link):
                    return out
            except Exception:
                continue
    except Exception:
        pass

    return out


def is_strict_address_candidate(text: str) -> bool:
    """Return True if text looks like a physical postal address.

    Heuristics: must contain a street number OR a postal code, and a street keyword
    (street type) OR common postal code pattern. This avoids grabbing hero text.
    """
    if not text or len(text.strip()) < 10:
        return False
    t = text.strip()
    # must have a street number or postal code
    has_number = bool(re.search(r"\d{1,5}", t))
    has_postal = bool(re.search(r"\b\d{5}(?:-\d{4})?\b", t)) or bool(re.search(r"\b[A-Z]{1,2}\d[A-Z\d]?\s*\d[A-Z]{2}\b", t, re.I))
    has_street = bool(re.search(STREET_KEYWORDS, t, re.I))

    # require (number and street) OR postal
    if (has_number and has_street) or has_postal:
        # also avoid lines that are too long and look like paragraphs
        if len(t) > 300:
            return False
        # avoid contact emails or long marketing lines
        if "@" in t or t.lower().startswith("news") or t.lower().startswith("data"):
            return False
        return True
    return False


def standardize_address_dict(raw: str) -> dict:
    out = {
        "STREET ADDRESS 1": "",
        "STREET ADDRESS 2": "",
        "CITY": "",
        "STATE": "",
        "PIN CODE": "",
        "COUNTRY": "",
    }
    if not raw:
        return out
    r = normalize_text(raw)
    for k, v in SHORT_FORMS.items():
        r = re.sub(rf"\b{k}\b", v, r, flags=re.I)

    parts = [p.strip() for p in re.split(r",|;|\n", r) if p.strip()]
    if parts:
        out["STREET ADDRESS 1"] = parts[0]
    if len(parts) >= 2:
        last = parts[-1]
        m = re.search(r"(\d{5}(?:-\d{4})?)", last)
        if m:
            out["PIN CODE"] = m.group(1)
            last = last.replace(m.group(1), "").strip()
        tokens = [t.strip() for t in last.split() if t.strip()]
        if tokens:
            if len(tokens) == 1 and len(tokens[0]) <= 3:
                out["STATE"] = tokens[0]
            else:
                out["COUNTRY"] = tokens[-1]
    if len(parts) >= 3:
        out["CITY"] = parts[-2]

    return out


def enrich_with_nominatim(record: dict) -> dict:
    q = ", ".join([record.get("STREET ADDRESS 1", ""), record.get("CITY", ""), record.get("STATE", ""), record.get("COUNTRY", "")])
    q = q.strip().strip(',')
    if not q:
        return record
    try:
        url = "https://nominatim.openstreetmap.org/search"
        res = requests.get(url, params={"q": q, "format": "json", "addressdetails": 1, "limit": 1}, headers={**HEADERS, "User-Agent": "SiteIntel/1.0 (mailto:you@example.com)"}, timeout=10)
        data = res.json()
        if data:
            addr = data[0].get("address", {})
            if not record.get("CITY") and addr.get("city"):
                record["CITY"] = addr.get("city").upper()
            if not record.get("STATE") and addr.get("state"):
                record["STATE"] = addr.get("state").upper()
            if not record.get("PIN CODE") and addr.get("postcode"):
                record["PIN CODE"] = addr.get("postcode")
            if not record.get("COUNTRY") and addr.get("country"):
                record["COUNTRY"] = addr.get("country").upper()
            time.sleep(1)
    except Exception:
        pass
    return record

# -------------------------------------------------
# ADDRESS EXTRACTION
# -------------------------------------------------
def extract_address(website: str) -> str:
    try:
        if not website.startswith("http"):
            website = "https://" + website

        res = requests.get(website, headers=HEADERS, timeout=10)
        soup = BeautifulSoup(res.text, "html.parser")

        tag = soup.find("address")
        if tag:
            import streamlit as st
            import pandas as pd
            import requests
            import re
            import time
            import hashlib
            from bs4 import BeautifulSoup
            from io import BytesIO
            from unidecode import unidecode
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            from openpyxl.worksheet.table import Table, TableStyleInfo
            from openpyxl.utils import get_column_letter

            # -------------------------------------------------
            # CONFIG
            # -------------------------------------------------
            HEADERS = {"User-Agent": "Mozilla/5.0"}
            REQUEST_DELAY = 0.15
            CRAWL_TIMEOUT = 6

            SHORT_FORMS = {
                "RD": "ROAD", "ST": "STREET", "AVE": "AVENUE",
                "BLVD": "BOULEVARD", "DR": "DRIVE", "LN": "LANE", "PL": "PLACE"
            }

            STANDARD_COUNTRIES = {
                "USA": "UNITED STATES OF AMERICA",
                "US": "UNITED STATES OF AMERICA",
            }

            STREET_KEYWORDS = r"\b(STREET|ST\.|ROAD|RD\.|AVE|AVENUE|BOULEVARD|BLVD|DR|DRIVE|LANE|LN|WAY|TERRACE|PLAZA|PL|COURT|CT)\b"

            # -------------------------------------------------
            # UTILITIES
            # -------------------------------------------------
            def ensure_scheme(url: str) -> str:
                if not url.startswith("http"):
                    return "https://" + url.lstrip("/")
                return url

            def normalize_text(text: str) -> str:
                if not text:
                    return ""
                text = unidecode(text).strip()
                return re.sub(r"\s+", " ", text)

            def hash_address(addr: dict) -> str:
                key = "|".join([addr.get(k, "") for k in ["STREET ADDRESS 1", "CITY", "STATE", "PIN CODE", "COUNTRY"]])
                return hashlib.md5(key.encode()).hexdigest()

            # -------------------------------------------------
            # PAGE SEARCH & EXTRACTION
            # -------------------------------------------------
            def candidate_paths():
                return ["/contact", "/contact-us", "/about", "/about-us", "/locations", "/location", "/contactus", "/company/contact"]

            def find_pages_from_home(home_url: str, max_pages=10):
                home = ensure_scheme(home_url)
                pages = [home]
                try:
                    r = requests.get(home, headers=HEADERS, timeout=CRAWL_TIMEOUT)
                    soup = BeautifulSoup(r.text, "html.parser")
                    for a in soup.find_all("a", href=True):
                        href = a["href"]
                        if href.startswith("/"):
                            pages.append(home.rstrip("/") + href)
                        elif href.startswith(home) or href.startswith("http") and home in href:
                            pages.append(href)
                        if len(set(pages)) >= max_pages:
                            break
                except Exception:
                    pass

                # add common paths
                base = re.sub(r"/+$", "", home)
                for p in candidate_paths():
                    pages.append(base + p)

                # dedupe
                out = []
                for p in pages:
                    if p not in out:
                        out.append(p)
                return out[:max_pages]

            def ddg_site_search(domain: str, query_terms="contact address", max_results=5):
                # Use DuckDuckGo HTML to search site pages for contact/address clues
                q = f"site:{domain} {query_terms}"
                try:
                    url = "https://html.duckduckgo.com/html/"
                    res = requests.post(url, data={"q": q}, headers=HEADERS, timeout=CRAWL_TIMEOUT)
                    soup = BeautifulSoup(res.text, "html.parser")
                    links = []
                    for a in soup.find_all("a", href=True):
                        href = a["href"]
                        if href.startswith("http") and domain in href:
                            links.append(href)
                        if len(links) >= max_results:
                            break
                    return links
                except Exception:
                    return []

            def find_address_in_html(text: str):
                if not text:
                    return None
                # Search for an address-like block using keywords and number patterns
                for line in text.splitlines():
                    line = line.strip()
                    if len(line) < 10:
                        continue
                    if re.search(STREET_KEYWORDS, line, re.I) and re.search(r"\d{1,5}", line):
                        return normalize_text(line)
                # fallback: search paragraphs
                m = re.search(r"([\w\s,.-]{10,150}" + STREET_KEYWORDS + r"[\w\s,.-]{0,150})", text, re.I)
                if m:
                    return normalize_text(m.group(0))
                return None

            def extract_address(website: str):
                # returns (raw_address_text, found_page_url)
                if not website:
                    return "", ""
                domain = re.sub(r"https?://", "", website).split("/", 1)[0]
                pages = find_pages_from_home(website, max_pages=12)

                # Try pages from site
                for p in pages:
                    try:
                        r = requests.get(ensure_scheme(p), headers=HEADERS, timeout=CRAWL_TIMEOUT)
                        soup = BeautifulSoup(r.text, "html.parser")
                        # check <address>
                        addr_tag = soup.find("address")
                        if addr_tag:
                            txt = addr_tag.get_text(" ", strip=True)
                            found = find_address_in_html(txt)
                            if found:
                                return found, ensure_scheme(p)

                        # check footer and contact sections
                        for selector in ["footer", "p", "div", "section"]:
                            for el in soup.find_all(selector):
                                txt = el.get_text(" ", strip=True)
                                if re.search(STREET_KEYWORDS, txt, re.I) and re.search(r"\d{1,5}", txt):
                                    return normalize_text(txt), ensure_scheme(p)
                    except Exception:
                        continue

                # If not found, try DDG site search for likely pages
                ddg_links = ddg_site_search(domain, query_terms="contact address", max_results=6)
                for link in ddg_links:
                    try:
                        r = requests.get(link, headers=HEADERS, timeout=CRAWL_TIMEOUT)
                        soup = BeautifulSoup(r.text, "html.parser")
                        txt = soup.get_text(" ", strip=True)
                        found = find_address_in_html(txt)
                        if found:
                            return found, link
                    except Exception:
                        continue

                return "", ""

            # -------------------------------------------------
            # PARSING / STANDARDIZATION
            # -------------------------------------------------
            def standardize_address(raw: str) -> dict:
                # parse raw address into components heuristically
                out = {
                    "STREET ADDRESS 1": "",
                    "STREET ADDRESS 2": "",
                    "CITY": "",
                    "STATE": "",
                    "PIN CODE": "",
                    "COUNTRY": "",
                }
                if not raw:
                    return out
                raw = normalize_text(raw)
                # replace short forms
                for k, v in SHORT_FORMS.items():
                    raw = re.sub(rf"\b{k}\b", v, raw, flags=re.I)

                # split by commas
                parts = [p.strip() for p in re.split(r",|;|\n", raw) if p.strip()]
                if parts:
                    out["STREET ADDRESS 1"] = parts[0]
                if len(parts) >= 2:
                    # try to assign city/state/pin from the last part
                    last = parts[-1]
                    # postal code
                    m = re.search(r"(\d{5}(?:-\d{4})?)", last)
                    if m:
                        out["PIN CODE"] = m.group(1)
                        last = last.replace(m.group(1), "").strip()
                    # state (2-letter) or full
                    tokens = [t.strip() for t in last.split() if t.strip()]
                    if tokens:
                        if len(tokens) == 1 and len(tokens[0]) <= 3:
                            out["STATE"] = tokens[0]
                        else:
                            out["COUNTRY"] = tokens[-1]

                # try to extract city from middle part
                if len(parts) >= 3:
                    out["CITY"] = parts[-2]

                return out

            # -------------------------------------------------
            # FREE NOMINATIM ENRICHMENT (optional, free)
            # -------------------------------------------------
            def enrich_with_nominatim(record: dict) -> dict:
                # Only call when we have at least street or city
                q = ", ".join([record.get("STREET ADDRESS 1", ""), record.get("CITY", ""), record.get("STATE", ""), record.get("COUNTRY", "")])
                q = q.strip().strip(",")
                if not q:
                    return record
                try:
                    url = "https://nominatim.openstreetmap.org/search"
                    res = requests.get(url, params={"q": q, "format": "json", "addressdetails": 1, "limit": 1}, headers={**HEADERS, "User-Agent": "SiteIntel/1.0 (mailto:you@example.com)"}, timeout=10)
                    data = res.json()
                    if data:
                        addr = data[0].get("address", {})
                        if not record.get("CITY") and addr.get("city"):
                            record["CITY"] = addr.get("city").upper()
                        if not record.get("STATE") and addr.get("state"):
                            record["STATE"] = addr.get("state").upper()
                        if not record.get("PIN CODE") and addr.get("postcode"):
                            record["PIN CODE"] = addr.get("postcode")
                        if not record.get("COUNTRY") and addr.get("country"):
                            record["COUNTRY"] = addr.get("country").upper()
                        time.sleep(1)
                except Exception:
                    pass
                return record

            # -------------------------------------------------
            # CONFIDENCE SCORE
            # -------------------------------------------------
            def calculate_confidence(addr: dict) -> int:
                score = 0
                if addr.get("STREET ADDRESS 1"):
                    score += 40
                if addr.get("CITY"):
                    score += 15
                if addr.get("STATE"):
                    score += 15
                if addr.get("PIN CODE"):
                    score += 15
                if addr.get("COUNTRY"):
                    score += 15
                return min(score, 100)

            # -------------------------------------------------
            # EXCEL EXPORT
            # -------------------------------------------------
            def generate_excel(data: list) -> bytes:
                wb = Workbook()
                ws = wb.active
                ws.title = "SiteIntel Output"

                ws.merge_cells("A1:K1")
                ws["A1"] = "SiteIntel – By Kishor"
                ws["A1"].font = Font(size=16, bold=True)
                ws["A1"].alignment = Alignment(horizontal="center")

                headers = [
                    "STREET ADDRESS 1", "STREET ADDRESS 2",
                    "CITY", "STATE", "PIN CODE", "COUNTRY",
                    "CONFIDENCE SCORE", "DUPLICATE FLAG",
                    "MASTER RECORD ID", "DATA SOURCE LINK", "FOUND PAGE"
                ]

                ws.append(headers)
                for c in ws[2]:
                    c.font = Font(bold=True)

                for r in data:
                    ws.append([r.get(h, "") for h in headers])

                ref = f"A2:K{len(data)+2}"
                table = Table(displayName="AddressTable", ref=ref)
                table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
                ws.add_table(table)

                for col in range(1, len(headers)+1):
                    ws.column_dimensions[get_column_letter(col)].width = 25

                ws.freeze_panes = "A3"
                out = BytesIO()
                wb.save(out)
                return out.getvalue()

            # -------------------------------------------------
            # STREAMLIT UI
            # -------------------------------------------------
            st.set_page_config(page_title="SiteIntel – By Kishor", layout="wide", page_icon="📍")
            st.markdown("""
            <div style='display:flex;align-items:center;gap:12px'>
              <div style='font-size:36px'>📍</div>
              <div>
                <h1 style='margin:0'>SiteIntel</h1>
                <div style='color:gray'>Enterprise Address Intelligence — Semi-Agentic (free sources)</div>
              </div>
            </div>
            """, unsafe_allow_html=True)

            mode = st.radio("Mode", ["Batch (Excel upload)", "Single Company"])

            if mode == "Batch (Excel upload)":
                uploaded = st.file_uploader("Upload Excel with company websites", type=["xlsx", "xls"])
                if st.button("🚀 Process Batch"):
                    if not uploaded:
                        st.warning("Upload a file first.")
                        st.stop()

                    df = pd.read_excel(uploaded)

                    # reuse the website column finder
                    def find_website_column(df: pd.DataFrame):
                        for c in df.columns:
                            if re.search(r"web|site|url", str(c), re.I):
                                return c
                        for c in df.columns:
                            try:
                                s = df[c].astype(str).str.strip().fillna("")
                            except Exception:
                                continue
                            if s.str.startswith("http").any():
                                return c
                        for c in df.columns:
                            try:
                                s = df[c].astype(str).str.strip().fillna("")
                            except Exception:
                                continue
                            if s.str.startswith("www.").any():
                                return c
                        domain_re = re.compile(r"\w+\.\w+")
                        for c in df.columns:
                            try:
                                s = df[c].astype(str).str.strip().fillna("")
                            except Exception:
                                continue
                            if s.apply(lambda x: bool(domain_re.search(x))).any():
                                return c
                        return None

                    url_col = find_website_column(df)
                    if not url_col:
                        st.error("No website column found. Please include a column with company website URLs.")
                        st.stop()

                    progress = st.progress(0)
                    records = []
                    seen = {}

                    for i, site in enumerate(df[url_col].astype(str)):
                        raw, page = extract_address(site)
                        parsed = standardize_address(raw)
                        parsed["DATA SOURCE LINK"] = site
                        parsed["FOUND PAGE"] = page
                        parsed = enrich_with_nominatim(parsed)
                        parsed["CONFIDENCE SCORE"] = calculate_confidence(parsed)

                        h = hash_address(parsed)
                        if h in seen:
                            parsed["DUPLICATE FLAG"] = "YES"
                            parsed["MASTER RECORD ID"] = seen[h]
                        else:
                            parsed["DUPLICATE FLAG"] = "NO"
                            parsed["MASTER RECORD ID"] = h[:8]
                            seen[h] = h[:8]

                        records.append(parsed)
                        progress.progress((i + 1) / max(1, len(df)))

                    st.success(f"Processed {len(records)} records")
                    st.dataframe(pd.DataFrame(records), use_container_width=True)
                    st.download_button("📥 Download Excel", generate_excel(records), "siteintel_output.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            else:
                st.subheader("Single company check")
                name = st.text_input("Company name (optional)")
                website = st.text_input("Official website (e.g. example.com or https://example.com)")
                if st.button("🔎 Check Single Company"):
                    if not website:
                        st.warning("Enter a website URL first.")
                        st.stop()
                    raw, page = extract_address(website)
                    parsed = standardize_address(raw)
                    parsed["DATA SOURCE LINK"] = website
                    parsed["FOUND PAGE"] = page
                    parsed = enrich_with_nominatim(parsed)
                    parsed["CONFIDENCE SCORE"] = calculate_confidence(parsed)
                    st.write("**Found page:**", page or "(not found)")
                    st.write("**Raw extract:**", raw or "(no extract)")
                    st.table(pd.DataFrame([parsed]))
                    st.download_button("📥 Download Excel", generate_excel([parsed]), "siteintel_single.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    # close the outer try in the top-level extract_address
    except Exception:
        pass

    return ""

        # -------------------------------------------------
# -------------------------------------------------
# CONFIDENCE SCORE
# -------------------------------------------------
def calculate_confidence(addr: dict) -> int:
    score = 0

    if addr["STREET ADDRESS 1"]:
        score += 40
    if addr["CITY"]:
        score += 15
    if addr["STATE"]:
        score += 15
    if addr["PIN CODE"]:
        score += 15
    if addr["COUNTRY"]:
        score += 15

    return min(score, 100)

# -------------------------------------------------
# EXCEL EXPORT
# -------------------------------------------------
def generate_excel(data: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "SiteIntel Output"

    ws.merge_cells("A1:J1")
    ws["A1"] = "SiteIntel – By Kishor"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = [
        "STREET ADDRESS 1", "STREET ADDRESS 2",
        "CITY", "STATE", "PIN CODE", "COUNTRY",
        "CONFIDENCE SCORE", "DUPLICATE FLAG",
        "MASTER RECORD ID", "DATA SOURCE LINK"
    ]

    ws.append(headers)

    for c in ws[2]:
        c.font = Font(bold=True)

    for r in data:
        ws.append([r.get(h, "") for h in headers])

    table = Table(
        displayName="AddressTable",
        ref=f"A2:J{len(data)+2}"
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showRowStripes=True
    )
    ws.add_table(table)

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 25

    ws.freeze_panes = "A3"

    out = BytesIO()
    wb.save(out)
    return out.getvalue()

# -------------------------------------------------
# STREAMLIT UI
# -------------------------------------------------
st.set_page_config(page_title="SiteIntel – By Kishor", layout="wide", page_icon="📍")
st.title("📍 SiteIntel")
st.caption("Enterprise Address Intelligence")

uploaded = st.file_uploader("Upload Excel with company websites", type=["xlsx", "xls"])

prefer_hq = st.checkbox("Prefer HQ/Corporate addresses only (skip store/location pages)", value=True)
extract_multiple = st.checkbox("Extract multiple locations per company (include store/outlet pages)", value=False)

if st.button("🚀 Process"):

    if not uploaded:
        st.warning("Upload a file first.")
        st.stop()

    df = pd.read_excel(uploaded)

    def find_website_column(df: pd.DataFrame):
        # 1) Column name heuristics
        for c in df.columns:
            if re.search(r"web|site|url", str(c), re.I):
                return c

        # 2) Values starting with http
        for c in df.columns:
            try:
                s = df[c].astype(str).str.strip().fillna("")
            except Exception:
                continue
            if s.str.startswith("http").any():
                return c

        # 3) Values starting with www.
        for c in df.columns:
            try:
                s = df[c].astype(str).str.strip().fillna("")
            except Exception:
                continue
            if s.str.startswith("www.").any():
                return c

        # 4) Values that look like domains (e.g. example.com)
        domain_re = re.compile(r"\w+\.\w+")
        for c in df.columns:
            try:
                s = df[c].astype(str).str.strip().fillna("")
            except Exception:
                continue
            if s.apply(lambda x: bool(domain_re.search(x))).any():
                return c

        return None

    url_col = find_website_column(df)

    if not url_col:
        st.error("No website column found.")
        st.stop()

    progress = st.progress(0)
    records = []
    seen = {}

    for i, site in enumerate(df[url_col].astype(str)):
        if extract_multiple:
            candidates = extract_all_addresses_site(site, limit=12)
            if not candidates:
                # fallback to single
                candidates = [extract_address_site(site, prefer_hq=bool(prefer_hq))]
        else:
            candidates = [extract_address_site(site, prefer_hq=bool(prefer_hq))]

        for raw, page in candidates:
            parsed = standardize_address_dict(raw)
            parsed["DATA SOURCE LINK"] = site
            parsed["FOUND PAGE"] = page
            parsed = enrich_with_nominatim(parsed)
            parsed["CONFIDENCE SCORE"] = calculate_confidence(parsed)

            h = hash_address(parsed)
            if h in seen:
                parsed["DUPLICATE FLAG"] = "YES"
                parsed["MASTER RECORD ID"] = seen[h]
            else:
                parsed["DUPLICATE FLAG"] = "NO"
                parsed["MASTER RECORD ID"] = h[:8]
                seen[h] = h[:8]

            records.append(parsed)

        progress.progress((i + 1) / max(1, len(df)))

    st.success(f"Processed {len(records)} records")

    st.dataframe(pd.DataFrame(records), use_container_width=True)

    st.download_button(
        "📥 Download Excel",
        generate_excel(records),
        "siteintel_output.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
