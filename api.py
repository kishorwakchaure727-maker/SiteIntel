from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import requests
from bs4 import BeautifulSoup
from io import BytesIO, StringIO
from unidecode import unidecode
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from pydantic import BaseModel
from typing import List
import json
import os

# Import functions from app.py (you might need to refactor this)
# For now, I'll duplicate the functions

# -------------------------------
# Configurations
# -------------------------------
GOOGLE_MAPS_API_KEY = os.getenv("GOOGLE_MAPS_API_KEY", "YOUR_GOOGLE_MAPS_API_KEY")

short_forms = {
    "RD": "ROAD", "ST": "STREET", "AVE": "AVENUE", "BLVD": "BOULEVARD",
    "DR": "DRIVE", "LN": "LANE", "PL": "PLACE", "CT": "COURT", "PKWY": "PARKWAY", "SQ": "SQUARE"
}

standard_countries = {
    "USA": "UNITED STATES OF AMERICA", "US": "UNITED STATES OF AMERICA",
    "UNITED STATES": "UNITED STATES OF AMERICA",
    "UK": "UNITED KINGDOM OF GREAT BRITAIN AND NORTHERN IRELAND",
    "UNITED KINGDOM": "UNITED KINGDOM OF GREAT BRITAIN AND NORTHERN IRELAND",
    "CHINA": "CHINA", "RUSSIA": "RUSSIAN FEDERATION",
    "SOUTH KOREA": "KOREA (REPUBLIC OF)", "KOREA": "KOREA (REPUBLIC OF)"
}

us_states = {
    "ALABAMA": "AL", "ALASKA": "AK", "ARIZONA": "AZ", "CALIFORNIA": "CA",
    "NEW YORK": "NY", "TEXAS": "TX", "FLORIDA": "FL", "ILLINOIS": "IL"
}

# -------------------------------
# Functions
# -------------------------------
def extract_address(website):
    try:
        response = requests.get(website, timeout=10)
        soup = BeautifulSoup(response.text, 'html.parser')
        address_tag = soup.find('address')
        if address_tag:
            return address_tag.get_text(separator=",")
        text = soup.get_text()
        lines = text.split('\n')
        for line in lines:
            for keyword in ["Head Office", "Corporate Office", "Address"]:
                if keyword.lower() in line.lower():
                    return line.strip()
        return ""
    except Exception:
        return ""

def standardize_address(raw_address):
    address = unidecode(raw_address).upper()
    for short, full in short_forms.items():
        address = re.sub(rf"\b{short}\b", full, address)

    parts = [p.strip() for p in address.split(",")]
    street_1 = parts[0] if len(parts) > 0 else ""
    street_2 = parts[1] if len(parts) > 1 else ""
    city = parts[2] if len(parts) > 2 else ""
    state = parts[3] if len(parts) > 3 else ""
    pin_code = parts[4] if len(parts) > 4 else ""
    country = parts[5] if len(parts) > 5 else ""

    for key, value in standard_countries.items():
        if country.startswith(key):
            country = value

    if country == "UNITED STATES OF AMERICA" and state in us_states:
        state = us_states[state]

    return {
        "STREET ADDRESS 1": street_1,
        "STREET ADDRESS 2": street_2,
        "CITY": city,
        "STATE": state,
        "PIN CODE": pin_code,
        "COUNTRY": country
    }

def enrich_with_google_maps(address):
    query = f"{address['STREET ADDRESS 1']} {address['CITY']} {address['STATE']} {address['COUNTRY']}"
    url = f"https://maps.googleapis.com/maps/api/geocode/json?address={query}&key={GOOGLE_MAPS_API_KEY}"
    try:
        response = requests.get(url)
        data = response.json()
        if data['status'] == 'OK':
            components = data['results'][0]['address_components']
            for comp in components:
                if 'locality' in comp['types']:
                    address['CITY'] = comp['long_name'].upper()
                if 'administrative_area_level_1' in comp['types']:
                    state_name = comp['long_name'].upper()
                    address['STATE'] = us_states.get(state_name, state_name)
                if 'country' in comp['types']:
                    address['COUNTRY'] = standard_countries.get(comp['long_name'].upper(), comp['long_name'].upper())
                if 'postal_code' in comp['types']:
                    address['PIN CODE'] = comp['long_name']
    except Exception:
        pass
    return address

def generate_excel(address_list):
    wb = Workbook()
    ws = wb.active
    ws.title = "Standardized Addresses"

    ws.merge_cells('A1:G1')
    ws['A1'] = "SiteIntel – By Kishor"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ["STREET ADDRESS 1", "STREET ADDRESS 2", "CITY", "STATE", "PIN CODE", "COUNTRY", "DATA SOURCE LINK"]
    ws.append(headers)

    for col in range(1, len(headers)+1):
        cell = ws.cell(row=2, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for addr in address_list:
        ws.append([addr.get(h, "") for h in headers])

    ref = f"A2:G{len(address_list)+2}"
    table = Table(displayName="AddressTable", ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    ws.freeze_panes = "A3"

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

# FastAPI app
app = FastAPI(
    title="SiteIntel Agentic AI API",
    description="Autonomous AI tool for company address extraction and standardization",
    version="1.0.0"
)

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class CompanyRequest(BaseModel):
    name: str
    website: str

class BatchRequest(BaseModel):
    companies: List[CompanyRequest]

class ProcessingResponse(BaseModel):
    status: str
    message: str
    data: Optional[List[dict]] = None

@app.get("/")
async def root():
    """Root endpoint with API information"""
    return {
        "message": "SiteIntel Agentic AI API",
        "description": "Autonomous company address extraction and standardization",
        "endpoints": {
            "/process-company": "Process single company",
            "/process-batch": "Process multiple companies",
            "/webhook-process": "Automatic file processing",
            "/health": "API health check"
        }
    }

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "timestamp": "2026-01-24"}

@app.post("/process-company", response_model=ProcessingResponse)
async def process_single_company(request: CompanyRequest):
    """Process a single company - Agentic endpoint"""
    try:
        raw_address = extract_address(request.website)
        standardized = standardize_address(raw_address)
        enriched = enrich_with_google_maps(standardized)
        enriched["DATA SOURCE LINK"] = request.website

        return ProcessingResponse(
            status="success",
            message=f"Successfully processed {request.name}",
            data=[enriched]
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Processing failed: {str(e)}")

@app.post("/process-batch", response_model=ProcessingResponse)
async def process_batch_companies(request: BatchRequest):
    """Process multiple companies - Agentic endpoint"""
    try:
        results = []
        for company in request.companies:
            raw_address = extract_address(company.website)
            standardized = standardize_address(raw_address)
            enriched = enrich_with_google_maps(standardized)
            enriched["DATA SOURCE LINK"] = company.website
            results.append(enriched)

        return ProcessingResponse(
            status="success",
            message=f"Successfully processed {len(results)} companies",
            data=results
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Batch processing failed: {str(e)}")

@app.post("/webhook-process")
async def webhook_process(file: UploadFile = File(...)):
    """Webhook endpoint for automatic file processing - Fully agentic"""
    try:
        # Validate file type
        if not file.filename.endswith(('.csv', '.xlsx')):
            raise HTTPException(status_code=400, detail="Only CSV and Excel files are supported")

        # Read uploaded file
        content = await file.read()
        if file.filename.endswith('.csv'):
            df = pd.read_csv(StringIO(content.decode('utf-8')))
        else:
            df = pd.read_excel(BytesIO(content))

        # Validate required columns
        required_cols = ["COMPANY NAME", "OFFICIAL WEBSITE"]
        if not all(col in df.columns for col in required_cols):
            raise HTTPException(status_code=400, detail=f"File must contain columns: {', '.join(required_cols)}")

        companies = [{"name": row["COMPANY NAME"], "website": row["OFFICIAL WEBSITE"]} for _, row in df.iterrows()]

        results = []
        for company in companies:
            raw_address = extract_address(company["website"])
            standardized = standardize_address(raw_address)
            enriched = enrich_with_google_maps(standardized)
            enriched["DATA SOURCE LINK"] = company["website"]
            results.append(enriched)

        # Generate Excel
        excel_data = generate_excel(results)

        return StreamingResponse(
            BytesIO(excel_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={file.filename.rsplit('.', 1)[0]}_processed.xlsx"}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"File processing failed: {str(e)}")

@app.post("/agentic-process")
async def agentic_process(request: dict):
    """Advanced agentic endpoint that can handle various input formats and make decisions"""
    try:
        # This endpoint can be extended to handle different types of requests
        # and make autonomous decisions about processing

        input_type = request.get("type", "unknown")

        if input_type == "company_list":
            companies = request.get("data", [])
            batch_request = BatchRequest(companies=[CompanyRequest(**c) for c in companies])
            return await process_batch_companies(batch_request)

        elif input_type == "single_company":
            company_data = request.get("data", {})
            company_request = CompanyRequest(**company_data)
            return await process_single_company(company_request)

        else:
            return {"status": "error", "message": "Unknown input type"}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Agentic processing failed: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)